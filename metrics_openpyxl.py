# -*- coding: utf-8 -*-
"""
Created on Tue May  4 16:40:02 2021

@author: PSANKA1
"""
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
print("enter file name with extension .xlsx:")
p=str(input())
path =p
# load excel file
book = openpyxl.load_workbook(path)
book.create_sheet('Sheet1')
book.create_sheet('Sheet2')
sheets_list=book.sheetnames
# select the sheet
sheet = sheets_list[0]
ws1 = book['Sheet1']
ws2 = book['Sheet2']
#Global Unix Team
U=["<Unix team member names>"]
# AP Unix Team
A=["<AP Unix team member names>"]
# AP Windows Team
W=["<AP windows team member names>"]
n= sheet.max_row
m= sheet.max_column
print(n,m)
ws1.append(["Infrastructure Change ID","Operational Categorization Tier 1","Operational Categorization Tier 2","Summary","Change Class","Change Coordinator","Scheduled Start Date","Scheduled End Date","Impact","Priority","Status","Status Reason","Requestor","Server CI","Environment","SC"])
k=2
for i in range(1, n+1):
    cell= sheet.cell(row = i, column = 14)
    cell2=sheet.cell(row = i, column = 1)
    if((cell.value!=None) and ((cell2.value==None) or (cell2.value[0:3]=="CRQ"))):
        for j in range(1, m+1):
            cell3=sheet.cell(row = i, column = j)
            ws1.cell(row = k, column = j).value=cell3.value
        k=k+1
n= ws1.max_row
print(n)
i=2
while(i<=n):
    c=1
    for j in range(i+1, n+1):
        cell= ws1.cell(row = j, column = 1)
        if(cell.value!=None):
            cell1=ws1.cell(row = i, column = 16)
            cell1.value=c
            i=j
            break;
        else:
            for k in range(1,m+1):
                if(k==7 or k==8):
                    cell2= ws1.cell(row = i, column = k)
                    cell2.alignment = Alignment(horizontal='center')
                    cell2.alignment = Alignment(vertical='center')
                    cell2.number_format = "m/d/yy h:mm"
                    cell1= ws1.cell(row = j, column = k)
                    cell1.value=cell2.value
                    cell1.alignment = Alignment(horizontal='center')
                    cell1.alignment = Alignment(vertical='center')
                    cell1.number_format = "m/d/yy h:mm"
                elif(k!=14 and k!=15):
                    cell2= ws1.cell(row = i, column = k)
                    cell2.alignment = Alignment(horizontal='center')
                    cell2.alignment = Alignment(vertical='center')
                    cell1= ws1.cell(row = j, column = k)
                    cell1.value=cell2.value
                    cell1.alignment = Alignment(horizontal='center')
                    cell1.alignment = Alignment(vertical='center')
                else:
                    cell2= ws1.cell(row = i, column = k)
                    cell2.alignment = Alignment(horizontal='center')
                    cell2.alignment = Alignment(vertical='center')
                    cell1= ws1.cell(row = j, column = k)
                    cell1.alignment = Alignment(horizontal='center')
                    cell1.alignment = Alignment(vertical='center')
            c=c+1
    else:
        cell1=ws1.cell(row = i, column = 16)
        cell1.value=c
        break;
for i in range(1,m+1):
    cell= ws1.cell(row = 1, column = i)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
ws1.column_dimensions['A'].width = 23
ws1.column_dimensions['B'].width = 29
ws1.column_dimensions['C'].width = 29
ws1.column_dimensions['D'].width = 90
ws1.column_dimensions['E'].width = 16
ws1.column_dimensions['F'].width = 25
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 18
ws1.column_dimensions['I'].width = 21
ws1.column_dimensions['J'].width = 18
ws1.column_dimensions['K'].width = 21
ws1.column_dimensions['L'].width = 14
ws1.column_dimensions['M'].width = 11
ws1.column_dimensions['N'].width = 12
ws1.column_dimensions['O'].width = 12
ws2.append(["CRQ","INFO","CCO","SC","OS","REG"])
j=2
for i in range(2, n+1):
    cell1= ws1.cell(row = i, column = 16)
    cell2= ws1.cell(row = i, column = 11)
    if(cell1.value!=None and(cell2.value=="Closed" or cell2.value=="Completed")):
        cell= ws1.cell(row = i, column = 1)
        k=1
        ws2.cell(row = j, column = k).value=cell.value
        k=k+1
        cell1= ws1.cell(row = i, column = 4)
        ws2.cell(row = j, column = k).value=cell1.value
        k=k+1
        cell1= ws1.cell(row = i, column = 6)
        ws2.cell(row = j, column = k).value=cell1.value
        k=k+1
        cell1= ws1.cell(row = i, column = 16)
        ws2.cell(row = j, column = k).value=cell1.value
        j=j+1
n= ws2.max_row
print(n)
for i in range(2, n+1):
    cell= ws2.cell(row = i, column = 3)
    if(cell.value in U):
        ws2.cell(row = i, column = 5).value="Unix"
        if(cell.value in A):
            ws2.cell(row = i, column = 6).value="AP"
        else:
           ws2.cell(row = i, column = 6).value="NA"
    else:
        ws2.cell(row = i, column = 5).value="Windows"
        if(cell.value in W):
            ws2.cell(row = i, column = 6).value="AP"
        else:
            ws2.cell(row = i, column = 6).value="NA" 
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 90
ws2.column_dimensions['C'].width = 25
m=ws2.max_column
for i in range(1,m+1):
    cell= ws2.cell(row = 1, column = i)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
print("end")
book.save(path)
